import os
import re
import logging
import hashlib
from datetime import datetime, timedelta
from typing import Final, Optional

from aiogram import Bot, Dispatcher, types, F
from aiogram.types import (
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from aiogram.filters import CommandStart, Command
from aiogram.enums import ParseMode
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.client.default import DefaultBotProperties
from openpyxl import Workbook, load_workbook

TOKEN = "8164683944:AAFblJC8b6i_2_poEqb7qnMnLd0WElfgG6Q"
BOT_USERNAME: Final = '@MepiDonor_bot'
ADMIN_PASSWORD_HASH = "5d5b09f6dcb2d53a5fffc60c4ac0d55fabdf556069d6631545f42aa6e3500f2e"
EXCEL_FILE = "donor_data.xlsx"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

COLUMNS = [
    "ID",
    "Telegram ID",
    "ФИО",
    "Группа",
    "Категория",
    "Кол-во Гаврилова",
    "Кол-во ФМБА",
    "Сумма донаций",
    "Дата последней донации Гаврилова",
    "Дата последней донации ФМБА",
    "Контакты соцсети",
    "Телефон",
    "В регистре ДКМ",
    "Согласие на рассылку",
    "Дата регистрации",
    "Организатор",
    "DonorCoin"
]

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Доноры"
        ws.append(COLUMNS)
        wb.save(EXCEL_FILE)


def add_donor(data: dict):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    donor_id = ws.max_row

    ws.append([
        donor_id,
        data.get('telegram_id'),
        data.get('full_name'),
        data.get('group'),
        data.get('category'),
        0,
        0,
        0,
        "",
        "",
        "",
        data.get('phone'),
        False,
        data.get('mailing_consent', False),
        datetime.now().strftime("%Y-%m-%d"),
        data.get('is_organizer', False),
        0  # DonorCoin
    ])

    wb.save(EXCEL_FILE)


def get_donor_by_telegram_id(telegram_id: int) -> Optional[dict]:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == telegram_id:
            return {
                'id': row[0],
                'telegram_id': row[1],
                'full_name': row[2],
                'group': row[3],
                'category': row[4],
                'gavrilovo_count': row[5],
                'fmba_count': row[6],
                'total_amount': row[7],
                'last_gavrilovo': row[8],
                'last_fmba': row[9],
                'social_contacts': row[10],
                'phone': row[11],
                'dkm_member': row[12],
                'mailing_consent': row[13],
                'registered_at': row[14],
                'is_organizer': row[15],
                'donor_coin': row[16]
            }
    return None


def update_donor(telegram_id: int, data: dict) -> bool:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if row[1].value == telegram_id:
            for key, value in data.items():
                if key == 'full_name':
                    row[2].value = value
                elif key == 'group':
                    row[3].value = value
                elif key == 'category':
                    row[4].value = value
                elif key == 'phone':
                    row[11].value = value
                elif key == 'dkm_member':
                    row[12].value = value
                elif key == 'mailing_consent':
                    row[13].value = value
                elif key == 'is_organizer':
                    row[15].value = value
                elif key == 'gavrilovo_count':
                    row[5].value = value
                elif key == 'fmba_count':
                    row[6].value = value
                elif key == 'last_gavrilovo':
                    row[8].value = value
                elif key == 'last_fmba':
                    row[9].value = value
                elif key == 'social_contacts':
                    row[10].value = value
                elif key == 'total_amount':
                    row[7].value = value
                elif key == 'donor_coin':
                    row[16].value = value

            wb.save(EXCEL_FILE)
            return True

    return False


def get_all_donors():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    donors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        donors.append({
            'id': row[0],
            'telegram_id': row[1],
            'full_name': row[2],
            'group': row[3],
            'category': row[4],
            'gavrilovo_count': row[5],
            'fmba_count': row[6],
            'total_amount': row[7],
            'last_gavrilovo': row[8],
            'last_fmba': row[9],
            'social_contacts': row[10],
            'phone': row[11],
            'dkm_member': row[12],
            'mailing_consent': row[13],
            'registered_at': row[14],
            'is_organizer': row[15],
            'donor_coin': row[16]
        })

    return donors


def add_donation(telegram_id: int, center: str, date: str):
    donor = get_donor_by_telegram_id(telegram_id)
    if not donor:
        return False

    if "гаврилова" in center.lower():
        update_data = {
            'gavrilovo_count': donor['gavrilovo_count'] + 1,
            'last_gavrilovo': date,
            'total_amount': donor['total_amount'] + 1,
            'donor_coin': donor['donor_coin'] + 500
        }
    else:
        update_data = {
            'fmba_count': donor['fmba_count'] + 1,
            'last_fmba': date,
            'total_amount': donor['total_amount'] + 1,
            'donor_coin': donor['donor_coin'] + 500
        }

    return update_donor(telegram_id, update_data)


def get_donations_stats():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    total_gavrilovo = 0
    total_fmba = 0
    total_donors = ws.max_row - 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_gavrilovo += row[5] or 0
        total_fmba += row[6] or 0

    return {
        'total_donors': total_donors,
        'total_gavrilovo': total_gavrilovo,
        'total_fmba': total_fmba,
        'total_donations': total_gavrilovo + total_fmba
    }

class RegistrationStates(StatesGroup):
    phone = State()
    full_name = State()
    category = State()
    group = State()
    consent = State()
    mailing_consent = State()


class OrganizerRegistrationStates(StatesGroup):
    password = State()
    full_name = State()


class DonationDayStates(StatesGroup):
    select_day = State()
    confirm = State()


class QuestionStates(StatesGroup):
    input_question = State()


class OrganizerAnswerStates(StatesGroup):
    select_question = State()
    input_answer = State()


class MailingStates(StatesGroup):
    select_recipients = State()
    input_message = State()
    confirm = State()


class ShopStates(StatesGroup):
    browse = State()
    view_item = State()
    confirm_purchase = State()

def validate_full_name(full_name: str) -> bool:
    pattern = r'^[А-ЯЁ][а-яё]+\s[А-ЯЁ][а-яё]+(?:\s[А-ЯЁ][а-яё]+)?$'
    return re.fullmatch(pattern, full_name) is not None


async def is_user_registered(telegram_id: int) -> bool:
    return get_donor_by_telegram_id(telegram_id) is not None


async def is_organizer(telegram_id: int) -> bool:
    donor = get_donor_by_telegram_id(telegram_id)
    return donor and donor.get('is_organizer', False)


async def get_user_info(telegram_id: int) -> Optional[dict]:
    return get_donor_by_telegram_id(telegram_id)


async def get_upcoming_donation_days(for_mifi: bool = True) -> list:
    return [
        {'id': 1, 'date': (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"),
         'center': 'Гаврилова', 'external_link': ''},
        {'id': 2, 'date': (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d"),
         'center': 'ФМБА', 'external_link': ''}
    ]


async def get_donor_balance(telegram_id: int) -> int:
    donor = get_donor_by_telegram_id(telegram_id)
    return donor.get('donor_coin', 0) if donor else 0


async def show_personal_cabinet(message: types.Message):
    user_info = await get_user_info(message.from_user.id)
    if not user_info:
        await message.answer("❌ Не удалось загрузить ваши данные. Попробуйте позже.")
        return

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📅 Записаться на День Донора", callback_data="register_dd")],
            [InlineKeyboardButton(text="✏️ Мои данные", callback_data="my_data")],
            [InlineKeyboardButton(text="🩸 История донаций", callback_data="donation_history")],
            [InlineKeyboardButton(text="🛍️ Магазин мерча", callback_data="shop")],
            [InlineKeyboardButton(text="❓ Вопрос организатору", callback_data="ask_question")],
            [InlineKeyboardButton(text="ℹ️ Информация о донорстве", callback_data="donation_info")],
        ]
    )

    await message.answer(
        f"👤 <b>Личный кабинет</b>\n\n"
        f"<b>ФИО:</b> {user_info['full_name']}\n"
        f"<b>Категория:</b> {user_info['category']}\n"
        f"<b>Группа/статус:</b> {user_info['group'] or 'не указано'}\n"
        f"<b>Донаций в Гаврилова:</b> {user_info['gavrilovo_count']}\n"
        f"<b>Донаций в ФМБА:</b> {user_info['fmba_count']}\n"
        f"<b>Всего донаций:</b> {user_info['total_amount']}\n"
        f"<b>DonorCoin:</b> {user_info['donor_coin']}\n"
        f"<b>В регистре ДКМ:</b> {'Да' if user_info['dkm_member'] else 'Нет'}\n"
        f"<b>Согласие на рассылку:</b> {'Да' if user_info['mailing_consent'] else 'Нет'}\n\n"
        "Выбери действие:",
        reply_markup=kb
    )


async def show_organizer_panel(message: types.Message):
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📝 Создать День Донора", callback_data="create_dd")],
            [InlineKeyboardButton(text="📊 Статистика", callback_data="stats")],
            [InlineKeyboardButton(text="❓ Вопросы от доноров", callback_data="view_questions")],
            [InlineKeyboardButton(text="📢 Сделать рассылку", callback_data="make_mailing")],
            [InlineKeyboardButton(text="🛍️ Управление магазином", callback_data="manage_shop")],
            [InlineKeyboardButton(text="📤 Выгрузить данные", callback_data="export_data")]
        ]
    )

    await message.answer(
        "👨‍⚕️ <b>Панель организатора</b>\n\n"
        "Выберите действие:",
        reply_markup=kb
    )
@dp.message(CommandStart())
async def cmd_start(message: types.Message, state: FSMContext):
    if await is_organizer(message.from_user.id):
        await show_organizer_panel(message)
    elif await is_user_registered(message.from_user.id):
        await show_personal_cabinet(message)
    else:
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📱 Отправить номер", request_contact=True)],
                [KeyboardButton(text="👨‍⚕️ Я организатор")]
            ],
            resize_keyboard=True
        )
        await message.answer(
            "👋 Привет! Я бот для организации Дня Донора в НИЯУ МИФИ.\n\n"
            "Я помогу тебе:\n"
            "✅ Зарегистрироваться как донор\n"
            "📅 Записаться на День Донора\n"
            "📊 Следить за своей историей донаций\n"
            "🛍️ Покупать мерч за DonorCoin\n\n"
            "Для начала отправь мне свой номер телефона или выбери 'Я организатор':",
            reply_markup=kb
        )
        await state.set_state(RegistrationStates.phone)


@dp.message(RegistrationStates.phone, F.text == "👨‍⚕️ Я организатор")
async def organizer_auth(message: types.Message, state: FSMContext):
    await message.answer(
        "🔐 Введите пароль для доступа в панель организатора:",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.set_state(OrganizerRegistrationStates.password)


@dp.message(OrganizerRegistrationStates.password)
async def organizer_password_check(message: types.Message, state: FSMContext):
    password = message.text.strip()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    print(hashed_password)
    logger.info(f"Input password hash: {hashed_password}")

    if hashed_password != ADMIN_PASSWORD_HASH:
        await message.answer("❌ Неверный пароль. Попробуйте еще раз.")
        return

    await message.answer(
        "✅ Пароль верный. Теперь введите ваше ФИО:",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.set_state(OrganizerRegistrationStates.full_name)

@dp.message(OrganizerRegistrationStates.full_name)
async def organizer_registration_complete(message: types.Message, state: FSMContext):
    full_name = message.text.strip()

    if not validate_full_name(full_name):
        await message.answer(
            "❌ Неверный формат ФИО. Пожалуйста, введите в формате: Фамилия Имя Отчество"
        )
        return
    donor = get_donor_by_telegram_id(message.from_user.id)
    if donor:
        update_donor(message.from_user.id, {
            'full_name': full_name,
            'is_organizer': True,
            'mailing_consent': True,
            'dkm_member': True
        })
    else:
        add_donor({
            'telegram_id': message.from_user.id,
            'full_name': full_name,
            'phone': "organizer",
            'category': "организатор",
            'is_organizer': True,
            'mailing_consent': True,
            'dkm_member': True
        })

    await message.answer(
        "🎉 Вы успешно зарегистрированы как организатор!\n\n"
        f"<b>ФИО:</b> {full_name}\n\n"
        "Теперь вам доступна панель организатора.",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.clear()
    await show_organizer_panel(message)


@dp.message(RegistrationStates.phone, F.contact)
async def contact_received(message: types.Message, state: FSMContext):
    phone = message.contact.phone_number
    await state.update_data(phone=phone)
    await message.answer(
        "Спасибо! Теперь введи своё ФИО (Фамилия Имя Отчество):",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.set_state(RegistrationStates.full_name)


@dp.message(RegistrationStates.full_name)
async def full_name_received(message: types.Message, state: FSMContext):
    full_name = message.text.strip()

    if not validate_full_name(full_name):
        await message.answer(
            "❌ Неверный формат ФИО. Пожалуйста, введи своё полное имя в формате:\n"
            "<b>Фамилия Имя Отчество</b> (если есть) с заглавных букв.\n"
            "Пример: <i>Иванов Иван Иванович</i> или <i>Петрова Анна</i>"
        )
        return

    await state.update_data(full_name=full_name)

    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Студент")],
            [KeyboardButton(text="Сотрудник")],
            [KeyboardButton(text="Внешний донор")]
        ],
        resize_keyboard=True
    )

    await message.answer(
        "Выбери свою категорию:",
        reply_markup=kb
    )
    await state.set_state(RegistrationStates.category)


@dp.message(RegistrationStates.category)
async def category_received(message: types.Message, state: FSMContext):
    category = message.text.lower()
    valid_categories = ['студент', 'сотрудник', 'внешний донор']

    if category not in valid_categories:
        await message.answer("Пожалуйста, выбери одну из предложенных категорий.")
        return

    await state.update_data(category=category)

    if category == 'студент':
        await message.answer(
            "Введи номер своей учебной группы:",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.set_state(RegistrationStates.group)
    else:
        await state.update_data(group=None)
        await ask_for_consent(message, state)


@dp.message(RegistrationStates.group)
async def group_received(message: types.Message, state: FSMContext):
    group = message.text.strip()
    await state.update_data(group=group)
    await ask_for_consent(message, state)


async def ask_for_consent(message: types.Message, state: FSMContext):
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Даю согласие", callback_data="consent_yes")],
            [InlineKeyboardButton(text="❌ Не согласен", callback_data="consent_no")]
        ]
    )

    await message.answer(
        "❗ Перед завершением регистрации необходимо дать согласие:\n\n"
        "Я согласен на обработку моих персональных данных в соответствии с Федеральным законом №152-ФЗ.",
        reply_markup=kb
    )
    await state.set_state(RegistrationStates.consent)


@dp.callback_query(RegistrationStates.consent, F.data.startswith("consent_"))
async def consent_received(callback: types.CallbackQuery, state: FSMContext):
    if callback.data == "consent_no":
        await callback.message.answer(
            "❌ Регистрация отменена. Без согласия на обработку персональных данных "
            "мы не можем зарегистрировать вас как донора."
        )
        await state.clear()
        return

    await state.update_data(consent=True)

    data = await state.get_data()
    category = data.get('category', '')

    mailing_text = (
        "Хотели бы вы получать информационные рассылки о предстоящих Днях Донора "
        "и других мероприятиях, связанных с донорством?"
    )

    if category == 'внешний донор':
        mailing_text = (
            "Хотели бы вы получать информацию о специальных мероприятиях "
            "для внешних доноров и других возможностях сдать кровь?"
        )

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Да, согласен", callback_data="mailing_yes")],
            [InlineKeyboardButton(text="❌ Нет, не согласен", callback_data="mailing_no")]
        ]
    )

    await callback.message.answer(
        mailing_text,
        reply_markup=kb
    )
    await state.set_state(RegistrationStates.mailing_consent)
    await callback.answer()


@dp.callback_query(RegistrationStates.mailing_consent, F.data.startswith("mailing_"))
async def mailing_consent_received(callback: types.CallbackQuery, state: FSMContext):
    mailing_consent = callback.data == "mailing_yes"
    await state.update_data(mailing_consent=mailing_consent)

    data = await state.get_data()
    if get_donor_by_telegram_id(callback.from_user.id):
        await callback.message.answer("❌ Этот номер телефона уже зарегистрирован.")
        await state.clear()
        return

    add_donor({
        'telegram_id': callback.from_user.id,
        'phone': data['phone'],
        'full_name': data['full_name'],
        'category': data['category'],
        'group': data.get('group'),
        'mailing_consent': mailing_consent
    })

    mailing_status = "Да" if mailing_consent else "Нет"
    await callback.message.answer(
        "🎉 Регистрация успешно завершена!\n\n"
        f"<b>ФИО:</b> {data['full_name']}\n"
        f"<b>Категория:</b> {data['category']}\n"
        f"<b>Группа:</b> {data.get('group', 'не указана')}\n"
        f"<b>Согласие на рассылку:</b> {mailing_status}\n\n"
        "Теперь ты можешь записаться на День Донора или посмотреть свою информацию.",
        reply_markup=ReplyKeyboardRemove()
    )

    await show_personal_cabinet(callback.message)
    await state.clear()


@dp.callback_query(F.data == "my_data")
async def show_user_data(callback: types.CallbackQuery):
    user_info = await get_user_info(callback.from_user.id)
    if not user_info:
        await callback.message.answer("❌ Не удалось загрузить ваши данные.")
        return

    balance = await get_donor_balance(callback.from_user.id)

    await callback.message.answer(
        f"📋 <b>Мои данные</b>\n\n"
        f"<b>ФИО:</b> {user_info['full_name']}\n"
        f"<b>Категория:</b> {user_info['category']}\n"
        f"<b>Группа/статус:</b> {user_info['group'] or 'не указано'}\n"
        f"<b>Донаций в Гаврилова:</b> {user_info['gavrilovo_count']}\n"
        f"<b>Донаций в ФМБА:</b> {user_info['fmba_count']}\n"
        f"<b>Всего донаций:</b> {user_info['total_amount']}\n"
        f"<b>DonorCoin:</b> {balance}\n"
        f"<b>В регистре ДКМ:</b> {'Да' if user_info['dkm_member'] else 'Нет'}\n"
        f"<b>Согласие на рассылку:</b> {'Да' if user_info['mailing_consent'] else 'Нет'}\n\n"
        "Чтобы изменить данные, обратитесь к организатору."
    )
    await callback.answer()


@dp.callback_query(F.data == "donation_history")
async def show_donation_history(callback: types.CallbackQuery):
    user_info = await get_user_info(callback.from_user.id)
    if not user_info:
        await callback.message.answer("❌ Не удалось загрузить ваши данные.")
        return

    text = "🩸 <b>История донаций</b>\n\n"

    if user_info['last_gavrilovo']:
        text += (
            f"<b>Центр:</b> Гаврилова\n"
            f"<b>Количество донаций:</b> {user_info['gavrilovo_count']}\n"
            f"<b>Последняя донация:</b> {user_info['last_gavrilovo']}\n\n"
        )

    if user_info['last_fmba']:
        text += (
            f"<b>Центр:</b> ФМБА\n"
            f"<b>Количество донаций:</b> {user_info['fmba_count']}\n"
            f"<b>Последняя донация:</b> {user_info['last_fmba']}\n\n"
        )

    if not user_info['last_gavrilovo'] and not user_info['last_fmba']:
        text += "❌ У вас пока нет донаций."

    await callback.message.answer(text)
    await callback.answer()


@dp.callback_query(F.data == "donation_info")
async def show_donation_info(callback: types.CallbackQuery):
    text = (
        "ℹ️ <b>Информация о донорстве</b>\n\n"
        "🔹 <b>Перед сдачей крови:</b>\n"
        "- Выспитесь\n"
        "- Питайтесь правильно за 2 дня до донации\n"
        "- Не употребляйте алкоголь за 48 часов\n"
        "- Пейте больше воды\n\n"
        "🔹 <b>После сдачи крови:</b>\n"
        "- Отдохните 10-15 минут\n"
        "- Не курите 2 часа\n"
        "- Не снимайте повязку 3-4 часа\n"
        "- Избегайте физических нагрузок\n\n"
        "📌 Полная информация на сайте <a href='https://mephi.ru'>mephi.ru</a>"
    )
    await callback.message.answer(text, disable_web_page_preview=True)
    await callback.answer()


@dp.callback_query(F.data == "ask_question")
async def ask_question(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "✍️ Напишите ваш вопрос организаторам:",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.set_state(QuestionStates.input_question)
    await callback.answer()


@dp.message(QuestionStates.input_question)
async def question_received(message: types.Message, state: FSMContext):
    question = message.text.strip()

    if len(question) < 10:
        await message.answer("❌ Вопрос слишком короткий. Пожалуйста, уточните детали.")
        return

    user_info = await get_user_info(message.from_user.id)
    if not user_info:
        await message.answer("❌ Ошибка: не удалось загрузить ваши данные.")
        await state.clear()
        return

    await message.answer(
        "✅ Ваш вопрос отправлен организаторам. "
        "Мы пришлем ответ в этом чате, как только он будет готов."
    )
    await state.clear()


@dp.callback_query(F.data == "register_dd")
async def register_for_donation_day(callback: types.CallbackQuery, state: FSMContext):
    if not await is_user_registered(callback.from_user.id):
        await callback.message.answer("Сначала нужно зарегистрироваться. Нажми /start")
        await callback.answer()
        return

    user_info = await get_user_info(callback.from_user.id)
    if not user_info:
        await callback.message.answer("❌ Ошибка: не удалось загрузить ваши данные.")
        await callback.answer()
        return

    is_mifi = user_info['category'] in ['студент', 'сотрудник']
    days = await get_upcoming_donation_days(for_mifi=is_mifi)

    if not days:
        await callback.message.answer("❌ В ближайшее время нет запланированных Дней Донора.")
        await callback.answer()
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for day in days:
        kb.inline_keyboard.append([
            InlineKeyboardButton(
                text=f"{day['date']} ({day['center']})",
                callback_data=f"dd_day_{day['id']}"
            )
        ])

    await callback.message.answer(
        "📅 Выбери День Донора для записи:",
        reply_markup=kb
    )
    await state.set_state(DonationDayStates.select_day)
    await callback.answer()


@dp.callback_query(DonationDayStates.select_day, F.data.startswith("dd_day_"))
async def donation_day_selected(callback: types.CallbackQuery, state: FSMContext):
    day_id = int(callback.data.split("_")[2])
    await state.update_data(day_id=day_id)

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Подтвердить", callback_data="dd_confirm")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="dd_cancel")]
        ]
    )

    days = await get_upcoming_donation_days()
    selected_day = next((day for day in days if day['id'] == day_id), None)

    if not selected_day:
        await callback.message.answer("❌ Ошибка выбора дня. Попробуйте снова.")
        await state.clear()
        return

    await callback.message.answer(
        f"Ты выбрал День Донора:\n\n"
        f"<b>Дата:</b> {selected_day['date']}\n"
        f"<b>Центр:</b> {selected_day['center']}\n\n"
        "Подтверди запись:",
        reply_markup=kb
    )
    await state.set_state(DonationDayStates.confirm)
    await callback.answer()


@dp.callback_query(DonationDayStates.confirm, F.data == "dd_confirm")
async def donation_day_confirmed(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    day_id = data['day_id']

    days = await get_upcoming_donation_days()
    selected_day = next((day for day in days if day['id'] == day_id), None)

    if not selected_day:
        await callback.message.answer("❌ Ошибка: день донора не найден.")
        await state.clear()
        return

    if add_donation(callback.from_user.id, selected_day['center'], selected_day['date']):
        await callback.message.answer(
            "✅ Ты успешно записался на День Донора!\n\n"
            f"<b>Дата:</b> {selected_day['date']}\n"
            f"<b>Центр:</b> {selected_day['center']}\n\n"
            "Мы пришлем тебе напоминание за день до события."
        )
    else:
        await callback.message.answer("❌ Ошибка при записи. Попробуйте позже.")

    await state.clear()
    await callback.answer()


@dp.callback_query(F.data == "donation_confirmed")
async def confirm_donation(callback: types.CallbackQuery):
    donor_id = callback.from_user.id
    donor = get_donor_by_telegram_id(donor_id)

    if donor:
        new_balance = donor['donor_coin'] + 500
        update_donor(donor_id, {'donor_coin': new_balance})

        await callback.message.answer(
            "🎉 Спасибо за донацию! Вам начислено 500 DonorCoin!\n"
            "Вы можете потратить их в нашем магазине мерча."
        )
    else:
        await callback.message.answer("❌ Ошибка: пользователь не найден.")

    await callback.answer()


@dp.callback_query(F.data == "shop")
async def show_shop(callback: types.CallbackQuery, state: FSMContext):
    items = [
        {'id': 1, 'name': 'Футболка "Я донор"', 'price': 300},
        {'id': 2, 'name': 'Значок донора', 'price': 150},
        {'id': 3, 'name': 'Термос', 'price': 500},
        {'id': 4, 'name': 'Браслет', 'price': 200}
    ]

    balance = await get_donor_balance(callback.from_user.id)

    kb = InlineKeyboardMarkup(inline_keyboard=[])
    for item in items:
        kb.inline_keyboard.append([
            InlineKeyboardButton(
                text=f"{item['name']} - {item['price']} DonorCoin",
                callback_data=f"item_{item['id']}"
            )
        ])

    kb.inline_keyboard.append([
        InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_cabinet")
    ])

    await callback.message.answer(
        f"🛍️ <b>Магазин мерча</b>\n\n"
        f"💰 Ваш баланс: <b>{balance} DonorCoin</b>\n\n"
        "Выберите товар:",
        reply_markup=kb
    )
    await state.set_state(ShopStates.browse)
    await callback.answer()


@dp.callback_query(ShopStates.browse, F.data.startswith("item_"))
async def view_shop_item(callback: types.CallbackQuery, state: FSMContext):
    item_id = int(callback.data.split("_")[1])

    items = {
        1: {'name': 'Футболка "Я донор"', 'description': 'Стильная хлопковая футболка', 'price': 300},
        2: {'name': 'Значок донора', 'description': 'Металлический значок', 'price': 150},
        3: {'name': 'Термос', 'description': 'Термос с логотипом донорства', 'price': 500},
        4: {'name': 'Браслет', 'description': 'Силиконовый браслет', 'price': 200}
    }

    item = items.get(item_id)
    if not item:
        await callback.message.answer("❌ Товар не найден.")
        await callback.answer()
        return

    balance = await get_donor_balance(callback.from_user.id)

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🛒 Купить", callback_data=f"buy_{item_id}")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_shop")]
    ])

    await callback.message.answer(
        f"🛍️ <b>{item['name']}</b>\n\n"
        f"📝 Описание: {item['description']}\n"
        f"💰 Цена: <b>{item['price']} DonorCoin</b>\n"
        f"💳 Ваш баланс: <b>{balance} DonorCoin</b>",
        reply_markup=kb
    )
    await state.set_state(ShopStates.view_item)
    await state.update_data(item_id=item_id, item_price=item['price'])
    await callback.answer()


@dp.callback_query(ShopStates.view_item, F.data.startswith("buy_"))
async def confirm_purchase(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    item_id = data['item_id']
    item_price = data['item_price']
    balance = await get_donor_balance(callback.from_user.id)

    if balance < item_price:
        await callback.message.answer(
            "❌ Недостаточно DonorCoin для покупки.\n"
            f"💰 Ваш баланс: {balance} | Цена товара: {item_price}"
        )
        await callback.answer()
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Подтвердить", callback_data=f"confirm_buy_{item_id}")],
        [InlineKeyboardButton(text="❌ Отмена", callback_data=f"item_{item_id}")]
    ])

    await callback.message.answer(
        "⚠️ <b>Подтвердите покупку</b>\n\n"
        f"Вы уверены, что хотите купить этот товар за {item_price} DonorCoin?",
        reply_markup=kb
    )
    await state.set_state(ShopStates.confirm_purchase)
    await callback.answer()


@dp.callback_query(ShopStates.confirm_purchase, F.data.startswith("confirm_buy_"))
async def process_purchase(callback: types.CallbackQuery, state: FSMContext):
    item_id = int(callback.data.split("_")[2])
    donor_id = callback.from_user.id

    donor = get_donor_by_telegram_id(donor_id)
    if not donor:
        await callback.message.answer("❌ Ошибка: пользователь не найден.")
        await state.clear()
        return

    items = {
        1: {'name': 'Футболка "Я донор"', 'price': 300},
        2: {'name': 'Значок донора', 'price': 150},
        3: {'name': 'Термос', 'price': 500},
        4: {'name': 'Браслет', 'price': 200}
    }

    item = items.get(item_id)
    if not item:
        await callback.message.answer("❌ Товар не найден.")
        await state.clear()
        return

    if donor['donor_coin'] < item['price']:
        await callback.message.answer("❌ Недостаточно DonorCoin для покупки.")
        await state.clear()
        return
    new_balance = donor['donor_coin'] - item['price']
    update_donor(donor_id, {'donor_coin': new_balance})

    await callback.message.answer(
        "🎉 Покупка успешно оформлена!\n\n"
        f"🛍️ Товар: <b>{item['name']}</b>\n"
        f"💰 Списано: <b>{item['price']} DonorCoin</b>\n\n"
        "Организаторы свяжутся с вами для уточнения деталей."
    )

    await state.clear()
    await callback.answer()


@dp.callback_query(F.data == "view_questions")
async def view_questions(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer(
        "❓ <b>Список вопросов от доноров</b>\n\n"
        "В текущей реализации вопросы не сохраняются. "
        "Для полной функциональности добавьте сохранение вопросов в Excel."
    )
    await callback.answer()


@dp.callback_query(F.data == "make_mailing")
async def make_mailing(callback: types.CallbackQuery, state: FSMContext):
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👨‍🎓 Для МИФИ", callback_data="mailing_mifi")],
            [InlineKeyboardButton(text="👥 Для внешних доноров", callback_data="mailing_external")],
            [InlineKeyboardButton(text="📢 Для всех", callback_data="mailing_all")]
        ]
    )

    await callback.message.answer(
        "👥 <b>Выберите целевую аудиторию рассылки:</b>",
        reply_markup=kb
    )
    await state.set_state(MailingStates.select_recipients)
    await callback.answer()


@dp.callback_query(MailingStates.select_recipients, F.data.startswith("mailing_"))
async def select_mailing_recipients(callback: types.CallbackQuery, state: FSMContext):
    mailing_type = callback.data.split("_")[1]
    await state.update_data(mailing_type=mailing_type)

    await callback.message.answer(
        "✍️ Введите текст рассылки:",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.set_state(MailingStates.input_message)
    await callback.answer()


@dp.message(MailingStates.input_message)
async def input_mailing_message(message: types.Message, state: FSMContext):
    mailing_text = message.text.strip()
    if len(mailing_text) < 10:
        await message.answer("❌ Текст рассылки слишком короткий. Пожалуйста, напишите более развернутое сообщение.")
        return

    await state.update_data(mailing_text=mailing_text)

    data = await state.get_data()
    mailing_type = data['mailing_type']

    recipient_type = ""
    if mailing_type == "mifi":
        recipient_type = "для МИФИ (студенты и сотрудники)"
    elif mailing_type == "external":
        recipient_type = "для внешних доноров"
    else:
        recipient_type = "для всех доноров"

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Подтвердить", callback_data="mailing_confirm")],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="mailing_cancel")]
        ]
    )

    await message.answer(
        f"📢 <b>Подтвердите рассылку</b>\n\n"
        f"<b>Получатели:</b> {recipient_type}\n\n"
        f"<b>Текст:</b>\n{mailing_text}",
        reply_markup=kb
    )
    await state.set_state(MailingStates.confirm)
    await state.update_data(mailing_text=mailing_text)


@dp.callback_query(MailingStates.confirm, F.data == "mailing_confirm")
async def confirm_mailing(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    mailing_type = data['mailing_type']
    mailing_text = data['mailing_text']

    donors = get_all_donors()
    recipients = []

    for donor in donors:
        if not donor['mailing_consent']:
            continue

        if mailing_type == "mifi" and donor['category'] in ['студент', 'сотрудник']:
            recipients.append(donor['telegram_id'])
        elif mailing_type == "external" and donor['category'] == 'внешний донор':
            recipients.append(donor['telegram_id'])
        elif mailing_type == "all":
            recipients.append(donor['telegram_id'])

    success = 0
    failed = 0

    for recipient in recipients:
        try:
            await bot.send_message(
                recipient,
                f"📢 <b>Информационное сообщение</b>\n\n{mailing_text}"
            )
            success += 1
        except Exception as e:
            logger.error(f"Error sending mailing to {recipient}: {e}")
            failed += 1

    await callback.message.answer(
        f"✅ Рассылка завершена\n\n"
        f"Успешно отправлено: {success}\n"
        f"Не удалось отправить: {failed}"
    )
    await state.clear()
    await callback.answer()


@dp.callback_query(F.data == "create_dd")
async def create_donation_day(callback: types.CallbackQuery):
    await callback.message.answer(
        "Функционал создания Дня Донора будет реализован в следующей версии."
    )
    await callback.answer()


@dp.callback_query(F.data == "stats")
async def show_stats(callback: types.CallbackQuery):
    stats = get_donations_stats()
    donors = get_all_donors()

    students = sum(1 for d in donors if d['category'] == 'студент')
    staff = sum(1 for d in donors if d['category'] == 'сотрудник')
    external = sum(1 for d in donors if d['category'] == 'внешний донор')
    total_coins = sum(d.get('donor_coin', 0) for d in donors)

    await callback.message.answer(
        "📊 <b>Статистика</b>\n\n"
        f"<b>Всего доноров:</b> {stats['total_donors']}\n"
        f"<b>Студентов:</b> {students}\n"
        f"<b>Сотрудников:</b> {staff}\n"
        f"<b>Внешних доноров:</b> {external}\n\n"
        f"<b>Всего донаций в Гаврилова:</b> {stats['total_gavrilovo']}\n"
        f"<b>Всего донаций в ФМБА:</b> {stats['total_fmba']}\n"
        f"<b>Всего донаций:</b> {stats['total_donations']}\n\n"
        f"<b>Всего DonorCoin в системе:</b> {total_coins}"
    )
    await callback.answer()


@dp.callback_query(F.data == "manage_shop")
async def manage_shop(callback: types.CallbackQuery):
    await callback.message.answer(
        "Функционал управления магазином будет реализован в следующей версии."
    )
    await callback.answer()


@dp.callback_query(F.data == "export_data")
async def export_data(callback: types.CallbackQuery):
    try:
        with open(EXCEL_FILE, 'rb') as file:
            await bot.send_document(
                chat_id=callback.from_user.id,
                document=types.BufferedInputFile(
                    file.read(),
                    filename="donor_data.xlsx"
                ),
                caption="📊 Текущие данные доноров"
            )
    except Exception as e:
        logger.error(f"Error exporting data: {e}")
        await callback.message.answer("❌ Ошибка при экспорте данных.")

    await callback.answer()


@dp.callback_query(F.data.in_(["back_to_shop", "back_to_cabinet"]))
async def handle_back_buttons(callback: types.CallbackQuery, state: FSMContext):
    try:
        if callback.data == "back_to_shop":
            await show_shop(callback, state)
        else:
            await state.clear()
            await callback.message.delete()
            await show_personal_cabinet(callback.message)
    except Exception as e:
        logger.error(f"Error handling back button: {e}")
        await callback.answer("Произошла ошибка, попробуйте еще раз")
    await callback.answer()


@dp.callback_query(ShopStates.browse, F.data == "back_to_cabinet")
@dp.callback_query(ShopStates.view_item, F.data == "back_to_cabinet")
@dp.callback_query(ShopStates.confirm_purchase, F.data == "back_to_cabinet")
async def back_to_cabinet_from_shop(callback: types.CallbackQuery, state: FSMContext):
    try:
        await state.clear()
        await callback.message.delete()
        await show_personal_cabinet(callback.message)
    except Exception as e:
        logger.error(f"Error returning to cabinet: {e}")
        await callback.answer("Произошла ошибка, попробуйте еще раз")
    await callback.answer()


@dp.message(Command("cabinet"))
async def cmd_cabinet(message: types.Message):
    if await is_organizer(message.from_user.id):
        await show_organizer_panel(message)
    elif not await is_user_registered(message.from_user.id):
        await message.answer("Сначала нужно зарегистрироваться. Нажми /start")
    else:
        await show_personal_cabinet(message)


if __name__ == "__main__":
    import asyncio
    init_excel()

    logger.info("Starting bot...")
    asyncio.run(dp.start_polling(bot))
